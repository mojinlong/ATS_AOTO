#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author   : Spw
# @Time     : 2021/5/19 15:49
# @File     : do_excel.py
# @Project  : integration-tests-insight
# -*- coding:utf-8 _*-

import openpyxl, os, pprint, pandas
from test_config import yamlconfig

testCase_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'test_data', 'test_case.xlsx')

cccccc = {
    'test_case': {
        'System': ['判断是否已登录', '查询所有字典数据', '查询某个字典类别', '根据字典类型查询某个字典类别', '查询所有字典类别', '查询所有字典类别(返回id和name)',
                   '查询所有字典类别(返回name和value)', '获取所有省份', '根据areaCode获取下级区域', '根据areaCode获取同级区域', '按顺序返回areaCode父层区域路径',
                   '一键审批', '获取图形验证码', '验证图形验证码是否正确', '查询所有2012类别', '查询所有2018类别', '配置列表，可根据查询条件', 'config 可选择字段',
                   '自动生成加工单', '测试触发生成计划', '自动生成结算单', '自动波次', '自动生成历史库存', '自动检查部门退货', '加载处理失败的消息', '重新发送失败的消息',
                   '加载打印信息', '批量加载打印信息', '标记打印成功'],
        'CreateDepartments': ['新增科室', '更新科室信息', '批量绑定商品和科室', '移除科室商品', '删除科室'],
        'goodsRequest_01': ['新增普通请领', '查询普通请领列表', '查询普通请领明细', '修改普通请领', '请领单撤回', '请领审核', '请领复核', '根据id获取请领单信息',
                            '根据id和状态获取请领单信息'],
        'goodsRequest_02': ['删除请领单'],
        'DistributionUnit_orderUnit': ['新建配送单位', '更新配送单位', '设置默认配送单位', '清空默认配送单位', '删除配送单位', '配送单位列表查询',
                                       '新建订货单位', '更新订货单位', '设置默认订货单位', '清空默认订货单位', '删除订货单位', '订货单位列表查询'],
        'std95GoodsCategory_goodsQuantityUnits': ['获取所有分类', '获取根节点列表', '获取每个根节点下面的子列表', '获取商品计量单位信息列表'],
        'custodian_01': ['新增一级供应商', '更新一级供应商', '启用/禁用', '根据id获取一级供应商详情', '分页获取一级供应商列表', '获取可用的一级供应商', '设置一级供应商账期'],
        'consume_01': ['批量消耗', '扫码消耗', '查询物资（商品、套包）', '查询并消耗', '分页查询商品消耗记录', '导出商品消耗记录', '分页查询定数包消耗记录', '导出定数包消耗记录',
                       '分页查询手术套包消耗记录', '导出手术套包消耗记录', '查询消耗记录详情', '扫码反消耗', '查询商品反消耗记录', '查询定数包反消耗记录',
                       '查询手术套包反消耗记录', '导出商品反消耗记录', '导出定数包反消耗记录', '导出手术套包反消耗记录', '查询定数包反消耗明细', '查询手术套包反消耗明细'],
        'consume_02': ['批量扫码反消耗'],
        'invoice_sync': ['查询列表--待审核', '待审核导出', '查询列表--待验收', '获取所有开票企业', '待验收导出', '查询列表--待支付', '待支付导出', '查询列表--支付完成',
                         '支付完成导出', '查询列表--驳回', '驳回导出', '发票审核', '发票验收', '发票支付', '发票作废', '编辑电子发票',
                         '分页获取货票同行开票列表', '货票同行发票上传', '根据发票号码查询蓝票', '查询发票汇总信息', '货票同行原票待修改详情',
                         '货票同行原票修改', '查看转账凭证'],
        'invoice_finStates': ['分页获取销后结算开票列表', '销后结算发票上传',
                              '销后结算待修改详情', '销后结算修改'],
        'invoice_Manual': ['根据发票查询发票明细', '电子发票红冲', '手工发票红冲', '修改手工红冲发票']
    },
    'test_case_1': {
        'GetAudit': ['分页查询审计记录', '查询操作历史', '创建角色'],
        # 'role': ['创建角色']
    }
}

a = yamlconfig.timeid(file_yaml='csv_config.yaml')._get_yaml_element_info()


class DoExcel:

    def __init__(self, sheet_name=None, file_name=testCase_dir, csv_dir=None):
        if sheet_name:
            self.file_name = file_name
            self.sheet_name = sheet_name
        elif csv_dir:
            new_csv_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'test_data',
                                       csv_dir)
            self.csv_dir = csv_dir
            self.file_name = new_csv_dir

    def get_cases_do_csv(self):
        file = pandas.read_csv(self.file_name, encoding='GBK')
        data = file.values.tolist()
        if self.csv_dir.split('.')[0] in a.keys():
            group_data = self.csv_dir.split('.')[0]
            data1 = self.case_group_by_csv(data, group_data)
            # self.case_to_ddt(data1)
            return data1

    def get_cases(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        cases = []  # 所有的cases数据
        max_row = sheet.max_row  # 获取最大行
        max_column = sheet.max_column  # 获取最大列
        # 单循环
        # for r in range(2, max_row + 1):
        #     case = {}
        #     case['case_id'] = sheet.cell(row=r, column=1).value
        #     case['title'] = sheet.cell(row=r, column=2).value
        #     case['url'] = sheet.cell(row=r, column=3).value
        #     case['data'] = sheet.cell(row=r, column=1).value
        #     case['method'] = sheet.cell(row=r, column=5).value
        #     case['expected'] = sheet.cell(row=r, column=6).value
        #     cases.append(case)
        # 双循环
        for r in range(2, max_row + 1):  # 遍历行
            case = {}
            for j in range(1, max_column + 1):
                key = sheet.cell(row=1, column=j).value  # 遍历列
                case[key] = sheet.cell(row=r, column=j).value
            cases.append(case)  # 将一行数据放到列表

        # workbook.save(self.file_name)
        workbook.close()

        data1 = self.case_group_by(cases)

        self.case_to_ddt(data1)

        return data1

    def case_group_by_csv(self, cases, group_data):
        new_cases = {}
        for i in a[group_data].keys():
            new_cases[i] = []
            for x in cases:
                if x[1] in a[group_data][i]:
                    new_cases[i].append(x)
        return new_cases

    def case_group_by(self, cases):
        new_cases = {}
        for i in a.keys():
            new_cases[i] = []
            for x in cases:
                if x['story'] in a[i]:
                    new_cases[i].append(x)
        return new_cases

    def case_to_ddt(self, case_data):
        # 用例转数据驱动格式
        for x in case_data.keys():
            new_case_list = []
            for i in case_data[x]:
                case = []
                for j in i.keys():
                    case.append(i[j])
                new_case_list.append(case)
                # i['data'] = [i['name'],i['data']]
            case_data[x] = new_case_list

        return case_data

    def get1_cases(self):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        cases = []  # 所有的cases数据
        max_row = sheet.max_row  # 获取最大行
        max_column = sheet.max_column  # 获取最大列
        # 单循环
        # for r in range(2, max_row + 1):
        #     case = {}
        #     case['case_id'] = sheet.cell(row=r, column=1).value
        #     case['title'] = sheet.cell(row=r, column=2).value
        #     case['url'] = sheet.cell(row=r, column=3).value
        #     case['data'] = sheet.cell(row=r, column=1).value
        #     case['method'] = sheet.cell(row=r, column=5).value
        #     case['expected'] = sheet.cell(row=r, column=6).value
        #     cases.append(case)
        # 双循环
        for r in range(2, max_row + 1):  # 遍历行
            case = {}
            for j in range(1, max_column + 1):
                key = sheet.cell(row=1, column=j).value  # 遍历列
                # while sheet.cell(row=r, column=j).value is None:
                #     r -= 1
                case[key] = sheet.cell(row=r, column=j).value
            cases.append(case)  # 将一行数据放到列表

        # workbook.save(self.file_name)
        workbook.close()

        data1 = self.case_group_by(cases)

        self.case_to_ddt(data1)

        return cases

    def write_result(self, row, actual, result):
        workbook = openpyxl.load_workbook(self.file_name)
        sheet = workbook[self.sheet_name]
        sheet.cell(row=row, column=7).value = actual
        sheet.cell(row=row, column=8).value = result
        workbook.save(self.file_name)
        workbook.close()


if __name__ == '__main__':
    # do_excle = DoExcel('接口')
    # pprint.pprint(do_excle.get_cases())
    pprint.pprint(DoExcel(csv_dir='test_case.csv').get_cases_do_csv())
